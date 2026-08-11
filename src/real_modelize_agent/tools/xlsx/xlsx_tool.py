from __future__ import annotations

from pathlib import Path
from typing import Any

from langchain_core.tools import StructuredTool


def read_xlsx(
    path: str,
    sheet: str | None = None,
    max_rows: int | str = 15,
    all_sheets: bool | str = False,
) -> dict[str, Any]:
    """只读 Excel（.xlsx/.xlsm）预览：返回工作表清单、每表形状与表头 + 前几行样例。

    用于建模手/编程手快速了解赛题附件结构，不需要写 pandas 脚本。公式取缓存值
    （data_only=True）；cell 统一转字符串，空单元格返回 None。
    """
    try:
        import openpyxl
    except ImportError as exc:
        return {"ok": False, "error": f"openpyxl is not installed: {exc}"}

    target = Path(path)
    if not target.is_file():
        return {"ok": False, "error": f"file not found: {path}"}
    if target.suffix.lower() not in {".xlsx", ".xlsm"}:
        return {"ok": False, "error": f"unsupported file type: {target.suffix or '(none)'}"}
    try:
        max_preview = max(1, min(50, int(max_rows)))
    except (TypeError, ValueError):
        max_preview = 15
    include_all = _coerce_bool(all_sheets)

    try:
        workbook = openpyxl.load_workbook(target, read_only=True, data_only=True)
    except Exception as exc:
        return {"ok": False, "error": f"cannot open workbook: {type(exc).__name__}: {exc}"}

    try:
        sheets = workbook.sheetnames
        selected = sheet or (sheets[0] if sheets else "")
        if selected not in sheets:
            return {"ok": False, "path": str(target), "sheets": sheets,
                    "error": f"sheet not found: {selected}"}
        to_preview = sheets if include_all else [selected]
        previews: dict[str, Any] = {}
        for name in to_preview:
            worksheet = workbook[name]
            previews[name] = _preview_sheet(worksheet, max_preview)
        return {
            "ok": True,
            "path": str(target),
            "sheets": sheets,
            "previews": previews,
        }
    finally:
        try:
            workbook.close()
        except Exception:
            pass


def build_xlsx_read_tool() -> StructuredTool:
    return StructuredTool.from_function(
        name="XlsxReadTool",
        func=read_xlsx,
        description=(
            "Read-only preview of an .xlsx/.xlsm workbook: sheet list, dimensions, headers and "
            "first rows per sheet (formula values, cells stringified). Args: path (required), "
            "optional sheet (name), optional max_rows (1-50, default 15), optional all_sheets (bool). "
            "Use to inspect competition attachments or data tables without writing pandas code."
        ),
    )


def _preview_sheet(worksheet, max_rows: int) -> dict[str, Any]:
    shape = [int(worksheet.max_row or 0), int(worksheet.max_column or 0)]
    rows: list[list[Any]] = []
    for index, row in enumerate(worksheet.iter_rows(values_only=True)):
        if index >= max_rows:
            break
        rows.append([_cell_text(value) for value in row])
    headers = rows[0] if rows else []
    return {
        "shape": shape,
        "headers": headers,
        "rows": rows[:max_rows],
    }


def _cell_text(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    return str(value)


def _coerce_bool(value: bool | str) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() not in {"false", "0", "no", "off"}
